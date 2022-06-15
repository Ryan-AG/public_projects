import os
import sys
import shutil
import logging
import pandas as pd
from glob import glob
from pathlib import Path
from win32com import client
from openpyxl import load_workbook
from datetime import date, timedelta
import pyautogui as gui
import concurrent.futures

todays_date = date.today()

path_dict = {
    'mk32_template_folder_path': Path(r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\MK32_Report_Template.xlsx'),
    'customer_center_folder_path': Path(r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\Customer Center//'),
    'maxim_report_folder_path': Path(r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\MaximRecurringEFT report//'),
    'temp_csm_file_name': Path('temp_csm_file.xlsx'),
    'csm_temp_folder_path': Path(r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\temp_c3_folder//'),
    'report_output_path': Path(r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\Output//'),
    'fds_path': Path(r"C:\Users\grane\Desktop\mk32_EDD_EFT Report\FDS//"),
}

temp_csm_file = Path(
    f"{path_dict['csm_temp_folder_path']}\\{path_dict['temp_csm_file_name']}")


LOG_FORMAT = "%(levelname)s:%(asctime)s:%(message)s"
logging.basicConfig(
    filename="mk32_test.log",
    level=logging.DEBUG,
    format=LOG_FORMAT,
)
logger = logging.getLogger()


def get_monday(todays_date: object):
    """Finds the date for Monday of the current week."""

    day_index = todays_date.weekday()
    monday = todays_date if day_index == 0 else todays_date - \
        timedelta(days=day_index)
    return monday


def convert_csm_to_temp(path_to_file, destination_path):
    """Converts csv file to xlsx for simplicity later"""

    if not os.path.exists(path_dict['csm_temp_folder_path']):
        os.mkdir(path_dict['csm_temp_folder_path'])
    try:
        logger.debug('Converting CSM file...')
        read_file = pd.read_csv(path_to_file)
        read_file.to_excel(destination_path, index=False, header=True)
    except FileNotFoundError as e:
        logger.error(e)
        sys.exit('CSM file could not be converted.')
    else:
        logger.debug("CSM file was successfully converted")


def transfer_data(original_worksheet, destination_worksheet, max_row, max_col, report):
    """Transfers the contents of one worksheet to another"""

    # * MK30 transfer must begin at row 8 to avoid clashing with the template
    start = 8 if report.upper() == "MK30" else 1
    for row in range(start, max_row + 1):
        for column in range(1, max_col + 1):
            cell_value = original_worksheet.cell(row=row, column=column)
            destination_worksheet.cell(
                row=row, column=column).value = cell_value.value


def find_max(worksheet):
    rows = worksheet.max_row
    columns = worksheet.max_column
    return rows, columns


def close_workbooks(*args):
    workbook = [*args]
    for book in workbook:
        book.close()


class FilePath:
    """Generate and hold the file paths for various reports"""

    def __init__(self, folder_path, todays_date, report_type):
        self.folder_path = folder_path
        self.todays_date = todays_date
        self.monday_date = get_monday(todays_date)
        self.year = self.monday_date.year
        self.month = self.monday_date.strftime('%m')
        self.day = self.monday_date.strftime('%d')
        self.report_type = report_type.upper()

    def __repr__(self):
        return f"FilePath(folder_path=r'{self.folder_path}', " \
               f"todays_date={self.todays_date}, " \
               f"report_type={self.report_type})"

    def __str__(self):
        return f'''{self.report_type} report: 
        Date of Report: {self.monday_date}
        Workbook Path: r"{self.file_selection()}"'''

    def file_format(self):
        if self.report_type == 'CSM':
            return '-'
        elif self.report_type == 'MAXIM':
            return '_'
        else:
            return ''

    def date_string(self):
        """Takes the file_format() and creates a date string for file_selection().

        Returns:
            string: matches the required date format for specific file types.
        """
        string = f'{self.year}{self.file_format()}{self.month}{self.file_format()}{self.day}'
        return string

    def file_selection(self):
        """Creates a list of all files meeting criteria. Returns first index."""
        files = glob(
            f'{self.folder_path}\\{self.report_type}*{self.date_string()}*')
        return files[0]


def main():

    logger.info(f'PROGRAM STARTED BY {os.getlogin()}.')

    # Set path objects for load_workbook() to pull in file.
    csm = FilePath(
        path_dict['customer_center_folder_path'], todays_date, 'csm')
    maxim = FilePath(
        path_dict['maxim_report_folder_path'], todays_date, 'maxim')
    mk30 = FilePath(
        path_dict['customer_center_folder_path'], todays_date, 'mk30')

    convert_csm_to_temp(
        csm.file_selection(),
        temp_csm_file
    )

    # Using threads to load workbooks
    with concurrent.futures.ThreadPoolExecutor() as executer:

        mk32_template = executer.submit(
            load_workbook, path_dict['mk32_template_folder_path']
        ).result()
        csm_workbook = executer.submit(
            load_workbook, temp_csm_file
        ).result()
        mk30_workbook = executer.submit(
            load_workbook, mk30.file_selection()
        ).result()
        maxim_workbook = executer.submit(
            load_workbook, maxim.file_selection()
        ).result()

    # Load all destination worksheets
    csm_mk32_worksheet = mk32_template['CSM']
    mk30_mk32_worksheet = mk32_template['PIF']
    maxim_mk32_worksheet = mk32_template['MAX_EFT']

    # Load all current worksheets
    csm_wb_worksheet = csm_workbook.worksheets[0]
    mk30_wb_worksheet = mk30_workbook.worksheets[0]
    maxim_wb_worksheet = maxim_workbook.worksheets[0]

    # Find max rows and columns
    csm_max_row, csm_max_col = find_max(csm_wb_worksheet)
    mk30_max_row, mk30_max_col = find_max(mk30_wb_worksheet)
    maxim_max_row, maxim_max_col = find_max(maxim_wb_worksheet)

    # Lists created to feed thread map()
    list_dict = {
        'orig_wb_list': [csm_wb_worksheet, mk30_wb_worksheet, maxim_wb_worksheet],
        'dest_wb_list': [csm_mk32_worksheet, mk30_mk32_worksheet, maxim_mk32_worksheet],
        'max_row_list': [csm_max_row, mk30_max_row, maxim_max_row],
        'max_col_list': [csm_max_col, mk30_max_col, maxim_max_col],
        'report_type_list': ['csm', 'mk30', 'maxim']
    }

    # Using threads to transfer data
    with concurrent.futures.ThreadPoolExecutor() as executer:
        executer.map(transfer_data,
                     list_dict['orig_wb_list'],
                     list_dict['dest_wb_list'],
                     list_dict['max_row_list'],
                     list_dict['max_col_list'],
                     list_dict['report_type_list']
                     )

    mk30_mk32_worksheet.cell(
        row=1, column=3).value = todays_date.strftime("%m/%d/%Y")
    mk30_mk32_worksheet.cell(
        row=2, column=3).value = get_monday(todays_date).strftime("%m/%d/%Y")

    close_workbooks(
        mk32_template,
        csm_workbook,
        mk30_workbook,
        maxim_workbook
    )

    os.remove(temp_csm_file)

    # ! Never save over the template !
    new_report_name = f'MK32_EDD_EFT Report_{csm.monday_date}'
    new_xl_path = f'{path_dict["report_output_path"]}\\{new_report_name}.xlsx'
    mk32_template.save(new_xl_path)
    new_report_path = f'{path_dict["report_output_path"]}\\{new_report_name}.pdf'

    '''Create a PDF of the PIF and Percentages worksheets.'''
    try:
        xl_wb = client.Dispatch("Excel.Application")

        report_wb = xl_wb.Workbooks.Open(new_xl_path)

        report_wb.Worksheets(["PIF", "Percentages"]).Select()

        xl_type_pdf = 0
        xl_quality_standard = 0

        xl_wb.ActiveSheet.ExportAsFixedFormat(xl_type_pdf,
                                              os.path.join(
                                                  path_dict['report_output_path'], new_report_name),
                                              xl_quality_standard, True, True)

    except Exception as e:
        gui.alert(f'The PDF was unable to be created.')
        logger.error(e)

    finally:
        report_wb.Close(SaveChanges=False)
        # xl_wb.Quit

    # A copy of the PDF must be distributed via FDS.
    shutil.copy(new_report_path, path_dict['fds_path'])

    logger.info('PROGRAM FINISHED.')


if __name__ == '__main__':
    main()
