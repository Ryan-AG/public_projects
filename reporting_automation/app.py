import os
import sys
import shutil
import logging
import pandas as pd
import pyautogui as gui
import time
from glob import glob
from pathlib import Path
from datetime import date, timedelta
from win32com import client
from openpyxl import load_workbook

todays_date = date.today()
user = os.getlogin()

LOG_FORMAT = "%(levelname)s:%(asctime)s:%(message)s"
logging.basicConfig(
    filename="mk32_test.log",
    level=logging.DEBUG,
    format=LOG_FORMAT,
)
logger = logging.getLogger()
logger.info(f'PROGRAM STARTED BY {user}.')


# Path to template file that will be written to
mk32_template_folder_path = Path(
    r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\MK32_Report_Template.xlsx')

# This folder houses both MK30 and CSM001 reports.
customer_center_folder_path = Path(
    r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\Customer Center//')

maxim_report_folder_path = Path(
    r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\MaximRecurringEFT report//')

# The temp_csm_file file will be used instead of the csv file and deleted at the end if this script.
temp_csm_file_name = Path('temp_csm_file.xlsx')

csm_temp_folder_path = Path(
    r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\temp_c3_folder//')

temp_csm_file = Path(f'{csm_temp_folder_path}\{temp_csm_file_name}')

report_output_path = Path(
    r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\Output//')

fds_path = r"C:\Users\grane\Desktop\mk32_EDD_EFT Report\FDS//"

report_excel_archive = r'C:\Users\grane\Desktop\mk32_EDD_EFT Report\Output\Excel Archive//'


def get_monday(todays_date):
    """Finds the Monday of the current week. Required to select the correct files"""
    day_index = todays_date.weekday()
    monday = todays_date if day_index == 0 else todays_date - \
        timedelta(day=day_index)
    return monday


def convert_csm_to_temp(path_to_file, destination_path):
    """Converts csv file to xlsx for simplicity later"""
    start = time.perf_counter()
    if not os.path.exists(csm_temp_folder_path):
        os.mkdir(csm_temp_folder_path)
    else:
        pass
    read_file = pd.read_csv(path_to_file)
    read_file.to_excel(destination_path, index=False, header=True)
    finish = time.perf_counter()
    logger.debug(
        f"convert_csm_to_temp() completed in {round(finish-start, 2)} second(s).")


def transfer_data(original_worksheet, destination_worksheet, max_row, max_col, report):
    """Transfers the contents of one worksheet to another"""
    # MK30 transfer must begin at row 8 to avoid clashing with the template
    start = 8 if report.upper() == "MK30" else 1
    for i in range(start, max_row + 1):
        for j in range(1, max_col + 1):
            cell_value = original_worksheet.cell(row=i, column=j)
            destination_worksheet.cell(
                row=i, column=j).value = cell_value.value
    pass


def move_pdf(src, dst):
    """Takes src file and moves to dst"""
    shutil.copy(f'{src}.pdf', dst)


class FilePath:
    """Generate and hold the file paths for various reports"""

    def __init__(self, folder_path, todays_date, report_type):
        self.folder_path = folder_path
        self.todays_date = todays_date
        self.monday_date = get_monday(todays_date)
        self.year = self.monday_date.year
        self.month = self.monday_date.strftime('%m')
        self.day = self.monday_date.strftime('%d')
        self.report_type = report_type.upper()

    def __repr__(self):
        return f"FilePath(folder_path=r'{self.folder_path}', todays_date={self.todays_date}, report_type={self.report_type})"

    def __str__(self):
        return f'''{self.report_type} report: 
        Date of Report: {self.monday_date}
        Workbook Path: r"{self.file_selection()}"'''

    def file_format(self):
        if self.report_type == 'CSM':
            return '-'
        elif self.report_type == 'MAXIM':
            return '_'
        else:
            return ''

    def date_string(self):
        """Takes the file_format() and creates a date string for file_selection().

        Returns:
            string: matches the required date format for specific file types.
        """        
        string = f'{self.year}{self.file_format()}{self.month}{self.file_format()}{self.day}'
        return string

    def file_selection(self):
        """Creates a list of all files meeting criteria. Returns first index."""
        files = glob(
            f'{self.folder_path}\{self.report_type}*{self.date_string()}*')
        return files[0]


# Set path objects for load_workbook() to pull in file.
csm = FilePath(customer_center_folder_path, todays_date, 'csm')
maxim = FilePath(maxim_report_folder_path, todays_date, 'maxim')
mk30 = FilePath(customer_center_folder_path, todays_date, 'mk30')


def main():
    print("Converting csm")
    # Create the temporary file that CSM data will be read from.
    convert_csm_to_temp(
        csm.file_selection(),
        temp_csm_file
    )
    print("CSM conversion complete.")

    # Load destination workbook and worksheets that will be written to.
    mk32_template = load_workbook(mk32_template_folder_path)
    csm_mk32_worksheet = mk32_template['CSM']
    mk30_mk32_worksheet = mk32_template['PIF']
    maxim_mk32_worksheet = mk32_template['MAX_EFT']

    # Load all current workbooks
    start = time.perf_counter()
    csm_workbook = load_workbook(temp_csm_file)
    finish = time.perf_counter()
    logger.debug(
        f"csm_workbook successfully loaded in {round(finish-start, 2)} second(s).")

    start = time.perf_counter()
    mk30_workbook = load_workbook(mk30.file_selection())
    finish = time.perf_counter()
    logger.debug(
        f"mk30_workbook successfully loaded in {round(finish-start, 2)} second(s).")

    start = time.perf_counter()
    maxim_workbook = load_workbook(maxim.file_selection())
    finish = time.perf_counter()
    logger.debug(
        f"maxim_workbook successfully loaded in {round(finish-start, 2)} second(s).")

    # Load all current worksheets
    csm_wb_worksheet = csm_workbook.worksheets[0]
    mk30_wb_worksheet = mk30_workbook.worksheets[0]
    maxim_wb_worksheet = maxim_workbook.worksheets[0]

    # Find max rows and columns
    csm_max_row = csm_wb_worksheet.max_row
    csm_max_col = csm_wb_worksheet.max_column

    mk30_max_row = mk30_wb_worksheet.max_row
    mk30_max_col = mk30_wb_worksheet.max_column

    maxim_max_row = maxim_wb_worksheet.max_row
    maxim_max_col = maxim_wb_worksheet.max_column

    # Transfer all data to template
    start = time.perf_counter()
    transfer_data(
        csm_wb_worksheet,
        csm_mk32_worksheet,
        csm_max_row,
        csm_max_col,
        "csm"
    )
    finish = time.perf_counter()
    logger.debug(
        f"transfer_data(csm) finished in {round(finish-start,2)} second(s)")

    start = time.perf_counter()
    transfer_data(
        mk30_wb_worksheet,
        mk30_mk32_worksheet,
        mk30_max_row,
        mk30_max_col,
        report="mk30"
    )
    finish = time.perf_counter()
    logger.debug(
        f"transfer_data(mk30) finished in {round(finish-start,2)} second(s)")

    start = time.perf_counter()
    transfer_data(
        maxim_wb_worksheet,
        maxim_mk32_worksheet,
        maxim_max_row,
        maxim_max_col,
        "maxim"
    )
    finish = time.perf_counter()
    logger.debug(
        f"transfer_data(maxim) finished in {round(finish-start,2)} second(s)")

    # Change the report date and generation date on the PIF worksheet
    mk30_mk32_worksheet.cell(
        row=1, column=3).value = todays_date.strftime("%m/%d/%Y")
    mk30_mk32_worksheet.cell(
        row=2, column=3).value = get_monday(todays_date).strftime("%m/%d/%Y")

    mk32_template.close()
    csm_workbook.close()
    mk30_workbook.close()
    maxim_workbook.close()

    os.remove(temp_csm_file)

    # Never save over the template!
    new_report_name = f'MK32_EDD_EFT Report_{csm.monday_date}'
    new_xl_path = f'{report_output_path}\\{new_report_name}.xlsx'
    mk32_template.save(new_xl_path)
    new_report_path = f'{report_output_path}\\{new_report_name}'

    '''Create a PDF of the PIF and Percentages worksheets.'''
    try:
        xl_wb = client.Dispatch("Excel.Application")

        report_wb = xl_wb.Workbooks.Open(new_xl_path)

        report_wb.Worksheets(["PIF", "Percentages"]).Select()

        xl_type_pdf = 0
        xl_quality_standard = 0

        xl_wb.ActiveSheet.ExportAsFixedFormat(xl_type_pdf,
                                              os.path.join(
                                                  report_output_path, new_report_name),
                                              xl_quality_standard, True, True)

    except Exception as e:
        gui.alert(f'''The PDF was unable to be created.
        Reason: {e}''')
        # gui.alert(e)

    finally:
        report_wb.Close(SaveChanges=False)
        xl_wb.Quit

        report_wb = None
        xl_wb = None

    logger.info('PROGRAM FINISHED.')

    # A copy of the PDF must be distributed via FDS.
    move_pdf(new_report_path, fds_path)


if __name__ == '__main__':
    main()
